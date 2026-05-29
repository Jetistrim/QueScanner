from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd  # type: ignore[reportMissingImports]
from openpyxl import load_workbook  # type: ignore[reportMissingImports]
from openpyxl.utils.cell import get_column_letter  # type: ignore[reportMissingImports]
from openpyxl.worksheet.table import Table, TableStyleInfo  # type: ignore[reportMissingImports]

from src.models import COLUMN_ORDER, validate_output_records


RowValue = str | int | float | None
RowData = dict[str, RowValue]


def build_dataframe(records: list[RowData]) -> pd.DataFrame:
    validated_records = validate_output_records(records)
    frame = pd.DataFrame(validated_records)
    for column in COLUMN_ORDER:
        if column not in frame.columns:
            frame[column] = None
    return frame.loc[:, COLUMN_ORDER]


def apply_basic_table_format(output_path: Path, row_count: int, column_count: int) -> None:
    if row_count <= 0 or column_count <= 0:
        return

    workbook = load_workbook(output_path)
    worksheet = workbook.active
    if worksheet is None:
        raise ValueError("Nao foi possivel obter a planilha ativa para formatacao")

    last_column = get_column_letter(column_count)
    last_row = row_count + 1
    table_ref = f"A1:{last_column}{last_row}"

    table = Table(displayName="QueScannerTable", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)
    worksheet.freeze_panes = "A2"

    workbook.save(output_path)


def export_to_xlsx(frame: pd.DataFrame, output_path: Path) -> None:
    casted_frame: Any = frame
    casted_frame.to_excel(output_path, index=False)
