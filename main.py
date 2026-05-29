from __future__ import annotations

import logging
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, TypeAlias, cast

import pandas as pd  # type: ignore[reportMissingImports]
import pdfplumber  # type: ignore[reportMissingImports]
from openpyxl import load_workbook  # type: ignore[reportMissingImports]
from openpyxl.worksheet.table import Table, TableStyleInfo  # type: ignore[reportMissingImports]
from openpyxl.utils.cell import get_column_letter  # type: ignore[reportMissingImports]

from src.extractor import parse_pdf_lines
from src.models import COLUMN_ORDER, validate_output_records


LOGGER_NAME = "que_scanner"
PROJECT_ROOT = Path(__file__).resolve().parent
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "outputs"
DEFAULT_OUTPUT_BASENAME = "consolidado"
RowValue: TypeAlias = str | int | float | None
RowData: TypeAlias = dict[str, RowValue]


def configure_logging() -> logging.Logger:
    logs_dir = PROJECT_ROOT / "logs"
    logs_dir.mkdir(exist_ok=True)

    logger = logging.getLogger(LOGGER_NAME)
    logger.setLevel(logging.INFO)
    logger.handlers.clear()
    logger.propagate = False

    formatter = logging.Formatter("[%(asctime)s] [%(levelname)s] - %(message)s")

    file_handler = logging.FileHandler(logs_dir / "app.log", encoding="utf-8")
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)

    stream_handler = logging.StreamHandler()
    stream_handler.setFormatter(formatter)
    logger.addHandler(stream_handler)

    return logger


def extract_lines_from_pdf(pdf_path: Path) -> list[str]:
    lines: list[str] = []
    pdfplumber_module: Any = pdfplumber
    with pdfplumber_module.open(pdf_path) as pdf:
        for page in getattr(pdf, "pages", []):
            extractor = getattr(page, "extract_text", None)
            text_value = extractor(layout=False) if callable(extractor) else ""
            text = text_value if isinstance(text_value, str) else ""
            for line in text.splitlines():
                cleaned = line.strip()
                if cleaned:
                    lines.append(cleaned)
    return lines


def build_dataframe(records: list[RowData]) -> pd.DataFrame:
    validated_records = validate_output_records(cast(list[dict[str, object]], records))
    frame = pd.DataFrame(validated_records)
    for column in COLUMN_ORDER:
        if column not in frame.columns:
            frame[column] = None
    return frame.loc[:, COLUMN_ORDER]


def apply_basic_table_format(output_path: Path, row_count: int, column_count: int) -> None:
    if row_count <= 0 or column_count <= 0:
        return

    workbook = load_workbook(output_path)
    worksheet = workbook.active
    if worksheet is None:
        raise ValueError("Não foi possível obter a planilha ativa para formatação")

    last_column = get_column_letter(column_count)
    last_row = row_count + 1
    table_ref = f"A1:{last_column}{last_row}"

    table = Table(displayName="QueScannerTable", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)
    worksheet.freeze_panes = "A2"

    workbook.save(output_path)


def build_generated_output_filename() -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{DEFAULT_OUTPUT_BASENAME}_{timestamp}.xlsx"


def ensure_unique_output_path(output_path: Path) -> Path:
    if not output_path.exists():
        return output_path

    stem = output_path.stem
    suffix = output_path.suffix
    counter = 1
    while True:
        candidate = output_path.with_name(f"{stem}_{counter:02d}{suffix}")
        if not candidate.exists():
            return candidate
        counter += 1


def resolve_output_path(raw_output_path: Path) -> Path:
    is_xlsx_file = raw_output_path.suffix.lower() == ".xlsx"

    if not is_xlsx_file:
        output_dir = raw_output_path if raw_output_path.is_absolute() else (PROJECT_ROOT / raw_output_path).resolve()
        return output_dir / build_generated_output_filename()

    if raw_output_path.is_absolute():
        return raw_output_path

    # If only a filename is provided, keep test artifacts under outputs/.
    if raw_output_path.parent == Path("."):
        return DEFAULT_OUTPUT_DIR / raw_output_path.name

    return (PROJECT_ROOT / raw_output_path).resolve()


def parse_cli_args(argv: list[str], logger: logging.Logger) -> tuple[list[Path], Path] | None:
    if len(argv) < 3:
        logger.error(
            "Uso esperado: python main.py <input1.pdf> [input2.pdf ...] <output.xlsx|output_dir>"
        )
        return None

    input_paths = [Path(value) for value in argv[1:-1]]
    raw_output_arg = Path(argv[-1])
    output_path = resolve_output_path(raw_output_arg)

    missing_inputs = [path for path in input_paths if not path.exists()]
    if missing_inputs:
        for missing in missing_inputs:
            logger.error("Arquivo PDF não encontrado: %s", missing)
        return None

    invalid_inputs = [path for path in input_paths if not path.is_file() or path.suffix.lower() != ".pdf"]
    if invalid_inputs:
        for invalid in invalid_inputs:
            logger.error("Entrada inválida (esperado arquivo .pdf): %s", invalid)
        return None

    if raw_output_arg.suffix.lower() == ".xlsx":
        logger.info("Modo de saída: arquivo explícito (.xlsx)")
    else:
        logger.info("Modo de saída: pasta (nome de arquivo gerado automaticamente)")

    return input_paths, output_path


def assign_global_row_ids(records: list[RowData]) -> list[RowData]:
    reassigned_records: list[RowData] = []
    for index, record in enumerate(records, start=1):
        reassigned_record = dict(record)
        reassigned_record["ID"] = index
        reassigned_records.append(reassigned_record)
    return reassigned_records


def run_processing(input_paths: list[Path], output_path: Path, logger: logging.Logger) -> int:
    try:
        all_records: list[RowData] = []

        for index, input_path in enumerate(input_paths, start=1):
            logger.info("Iniciando processamento (%s/%s): %s", index, len(input_paths), input_path)
            lines = extract_lines_from_pdf(input_path)
            records = parse_pdf_lines(lines, input_path=input_path, logger=logger)
            all_records.extend(records)
            logger.info(
                "Arquivo consolidado (%s/%s): %s | linhas=%s",
                index,
                len(input_paths),
                input_path,
                len(records),
            )

        all_records = assign_global_row_ids(all_records)

        frame = build_dataframe(all_records)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        final_output_path = ensure_unique_output_path(output_path)
        if final_output_path != output_path:
            logger.warning(
                "Arquivo de saída já existe. Nome alternativo aplicado: %s",
                final_output_path,
            )

        cast(Any, frame).to_excel(final_output_path, index=False)
        apply_basic_table_format(final_output_path, row_count=len(frame), column_count=len(frame.columns))
        logger.info(
            "Processamento concluído com sucesso: %s | arquivos=%s | linhas=%s",
            final_output_path,
            len(input_paths),
            len(all_records),
        )
        return 0
    except (OSError, ValueError, KeyError) as exc:
        logger.exception("Falha ao processar arquivo: %s", exc)
        return 1


def main() -> int:
    logger = configure_logging()

    parsed = parse_cli_args(sys.argv, logger)
    if parsed is None:
        return 1

    input_paths, output_path = parsed
    return run_processing(input_paths, output_path, logger)


if __name__ == "__main__":
    raise SystemExit(main())